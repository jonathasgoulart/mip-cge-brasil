"""
Gera A_mrio_official_v7_0.npy
Substituição completa dos fluxos interestaduais (Opção B):
  - Substitui blocos diagonais (intra-UF) e off-diagonal (inter-UF)
  - Dados da MIIP SS (Tabela 3) da IIOAS_BRUF_2019
  - VBP real de cada UF extraído da aba Produção (Tabela 1)
  - Crosswalk 68→67 setores (fusão S41+S42 → M41)

Estimativa de tempo: ~5-10 min (arquivo 120 MB)
"""
import numpy as np
import openpyxl
import json
import sys
import time
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')
t0 = time.time()

# ── Configurações ────────────────────────────────────────────────────────────
BASE      = Path(r'C:\Users\jonat\Documents\MIP e CGE')
IIOAS     = BASE / 'data' / 'IIOAS_BRUF_2019.xlsx'
CW_FILE   = BASE / 'output' / 'crosswalk' / 'crosswalk_iioas_mrio67.json'
OUT_DIR   = BASE / 'output' / 'final'
OUT_DIR.mkdir(parents=True, exist_ok=True)

N_IIOAS   = 68   # setores IIOAS
N_MRIO    = 67   # setores MRIO
N_UFS     = 27   # regiões

# ── 1. Crosswalk ─────────────────────────────────────────────────────────────
print("[1/5] Carregando crosswalk 68→67...")
with open(CW_FILE, encoding='utf-8') as f:
    cw = json.load(f)

# iioas_to_mrio[ii_0idx] = mi_0idx
iioas_to_mrio = {}
for row in cw['crosswalk']:
    ii = row['iioas_idx'] - 1
    mi = row['mrio67_idx'] - 1
    if ii not in iioas_to_mrio:
        iioas_to_mrio[ii] = mi

# Conjuntos de IIOAS que mapeiam para o mesmo MRIO (para agregação)
# Aqui: S40 (idx39) e S41 (idx40) → M40 (idx40)
from collections import defaultdict
mrio_from_iioas = defaultdict(list)
for ii, mi in iioas_to_mrio.items():
    mrio_from_iioas[mi].append(ii)

print(f"  OK: {len(iioas_to_mrio)} IIOAS → {len(set(iioas_to_mrio.values()))} MRIO67")
print(f"  Setores fundidos: { {mi: iioas for mi, iioas in mrio_from_iioas.items() if len(iioas) > 1} }")

# ── 2. Abrir arquivo e extrair VBP de cada UF ─────────────────────────────
print(f"\n[2/5] Abrindo IIOAS e extraindo VBP das 27 UFs (aba Produção)...")
wb = openpyxl.load_workbook(IIOAS, read_only=True, data_only=True)
names = wb.sheetnames

ws_prod = wb[names[4]]   # Produção
ws_ss   = wb[names[6]]   # MIIP SS

# VBP[uf_idx, setor_iioas_idx] = soma de toda a linha de produção
VBP = np.zeros((N_UFS, N_IIOAS), dtype=np.float64)

for uf_idx in range(N_UFS):
    row_start = 6 + uf_idx * N_IIOAS + 1    # 1-indexed
    row_end   = row_start + N_IIOAS - 1

    for s_idx, row in enumerate(ws_prod.iter_rows(
            min_row=row_start, max_row=row_end,
            min_col=5, values_only=True)):          # col 5+ = dados
        total = 0.0
        for v in row:
            try:
                total += float(v)
            except:
                pass
        VBP[uf_idx, s_idx] = total

    elapsed = time.time() - t0
    uf_labels = ['RO','AC','AM','RR','PA','AP','TO','MA','PI','CE','RN','PB',
                 'PE','AL','SE','BA','MG','ES','RJ','SP','PR','SC','RS','MS','MT','GO','DF']
    print(f"  UF {uf_idx+1:02d}/{N_UFS} ({uf_labels[uf_idx]}): VBP total R$ {VBP[uf_idx].sum():,.0f} Mi  [{elapsed:.0f}s]")

print(f"\n  VBP extraído. Total BR: R$ {VBP.sum():,.0f} Mi")
np.save(OUT_DIR.parent / 'crosswalk' / 'vbp_iioas_all_ufs.npy', VBP)

# ── 3. Extrair MIIP SS completa e converter para coeficientes ────────────────
print(f"\n[3/5] Extraindo MIIP SS (27×27 blocos 68×68) e convertendo para coef...")
# Vamos ler linha por linha e for region no loop externo
# Indexação: linha uf*68 + s_origem → coluna uf*68 + s_destino

# Alocar Z_iioas[uf_orig, s_orig, uf_dest, s_dest] — muito grande para memória
# Usamos abordagem streaming: lemos bloco por bloco diretamente para A68
# A68[uf_orig*N_IIOAS+s_orig, uf_dest*N_IIOAS+s_dest] = Z / VBP[uf_dest, s_dest]

# Tamanho total: 1836×1836 = 3.37M elementos float64 = ~27 MB — cabe na memória
N_TOTAL_IIOAS = N_UFS * N_IIOAS  # 1836
A68 = np.zeros((N_TOTAL_IIOAS, N_TOTAL_IIOAS), dtype=np.float32)

# Ler toda a MIIP SS de uma vez (linha por linha)
HEADER_ROWS = 6
DATA_COLS_START = 5  # 1-indexed (cols 1-4 são identificadores)

for row_iioas_idx, row in enumerate(ws_ss.iter_rows(
        min_row=HEADER_ROWS+1, max_row=HEADER_ROWS+N_TOTAL_IIOAS,
        min_col=DATA_COLS_START, max_col=DATA_COLS_START+N_TOTAL_IIOAS-1,
        values_only=True)):

    uf_orig  = row_iioas_idx // N_IIOAS
    s_orig   = row_iioas_idx %  N_IIOAS

    if s_orig == 0 and uf_orig % 5 == 0:
        elapsed = time.time() - t0
        print(f"  Lendo UF {uf_orig+1:02d}/{N_UFS} ({uf_labels[uf_orig]})... [{elapsed:.0f}s]")

    for col_idx, v in enumerate(row):
        try:
            z = float(v)
        except:
            z = 0.0
        if z == 0.0:
            continue

        uf_dest = col_idx // N_IIOAS
        s_dest  = col_idx %  N_IIOAS
        vbp_dest = VBP[uf_dest, s_dest]

        if vbp_dest > 0:
            A68[row_iioas_idx, col_idx] = z / vbp_dest

elapsed = time.time() - t0
print(f"\n  MIIP SS lida. Tempo total até agora: {elapsed:.0f}s")
print(f"  A68 shape: {A68.shape}, max={A68.max():.4f}, zeros: {(A68==0).mean()*100:.1f}%")

wb.close()

# ── 4. Agregar 68→67 via crosswalk ───────────────────────────────────────────
print(f"\n[4/5] Agregando 68→67 setores via crosswalk...")
N_TOTAL_MRIO = N_UFS * N_MRIO  # 1809
A_v7 = np.zeros((N_TOTAL_MRIO, N_TOTAL_MRIO), dtype=np.float32)

for uf_r in range(N_UFS):
    for uf_s in range(N_UFS):
        # Bloco IIOAS [uf_r × uf_s]: shape 68×68
        r0, r1 = uf_r * N_IIOAS, (uf_r+1) * N_IIOAS
        s0, s1 = uf_s * N_IIOAS, (uf_s+1) * N_IIOAS
        bloco68 = A68[r0:r1, s0:s1]

        # Bloco MRIO destino: shape 67×67
        mr0, mr1 = uf_r * N_MRIO, (uf_r+1) * N_MRIO
        ms0, ms1 = uf_s * N_MRIO, (uf_s+1) * N_MRIO

        # Agregar linhas e colunas
        bloco67 = np.zeros((N_MRIO, N_MRIO), dtype=np.float64)
        for ii_row in range(N_IIOAS):
            mi_row = iioas_to_mrio[ii_row]
            for ii_col in range(N_IIOAS):
                mi_col = iioas_to_mrio[ii_col]
                bloco67[mi_row, mi_col] += bloco68[ii_row, ii_col]

        # Média para setores fundidos (S41+S42 → M41): dividir por 2
        for mi, iioas_list in mrio_from_iioas.items():
            if len(iioas_list) > 1:
                bloco67[mi, :] /= len(iioas_list)
                bloco67[:, mi] /= len(iioas_list)

        A_v7[mr0:mr1, ms0:ms1] = bloco67.astype(np.float32)

    if uf_r % 5 == 0:
        elapsed = time.time() - t0
        print(f"  Agregado UF {uf_r+1:02d}/{N_UFS}... [{elapsed:.0f}s]")

print(f"  A_v7 shape: {A_v7.shape}")
col_sums = A_v7.sum(axis=0)
print(f"  Soma de colunas: min={col_sums.min():.3f} max={col_sums.max():.3f}")
unstable = (col_sums >= 1).sum()
print(f"  Colunas instáveis (soma>=1): {unstable}")

# ── 5. Salvar ─────────────────────────────────────────────────────────────────
print(f"\n[5/5] Salvando A_mrio_official_v7_0.npy...")
out_path = OUT_DIR / 'A_mrio_official_v7_0.npy'
np.save(out_path, A_v7)

elapsed = time.time() - t0
size_mb = out_path.stat().st_size / 1024**2
print(f"  Salvo: {out_path}")
print(f"  Tamanho: {size_mb:.1f} MB")
print(f"  Tempo total: {elapsed:.0f}s ({elapsed/60:.1f} min)")

# Comparacao rapida de multiplicadores
from scipy.linalg import inv as scipy_inv

rj_idx = 18
s0, s1 = rj_idx*N_MRIO, (rj_idx+1)*N_MRIO

A_rj_v7  = A_v7[s0:s1, s0:s1].astype(np.float64)
A_mrio_v6 = np.load(OUT_DIR / 'A_mrio_official_v6_1.npy')[s0:s1, s0:s1].astype(np.float64)

I67 = np.eye(N_MRIO)
mult_v7 = np.linalg.inv(I67 - A_rj_v7).sum(axis=0)
mult_v6 = np.linalg.inv(I67 - A_mrio_v6).sum(axis=0)

print(f"\n=== Comparação Multiplicadores RJ ===")
print(f"  v6.1: média={mult_v6.mean():.3f}  max={mult_v6.max():.3f}")
print(f"  v7.0: média={mult_v7.mean():.3f}  max={mult_v7.max():.3f}")
print(f"  Var%: {((mult_v7.mean()-mult_v6.mean())/mult_v6.mean()*100):+.1f}%")
print("\n[OK] MRIO v7.0 gerada com sucesso!")
