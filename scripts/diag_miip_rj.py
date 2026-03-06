"""
Diagnóstico: verifica a estrutura da MIIP SS e da aba Produção para o bloco RJ.
Objetivo: entender como calcular corretamente o VBP e os coeficientes técnicos.
"""
import openpyxl, sys, numpy as np
sys.stdout.reconfigure(encoding='utf-8')

FILE = r'C:\Users\jonat\Documents\MIP e CGE\data\IIOAS_BRUF_2019.xlsx'
N_S = 68

print('Abrindo arquivo...')
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
names = wb.sheetnames
print('Abas:', names)

# ── Diagnóstico MIIP SS ──────────────────────────────────────────────────────
ws = wb[names[6]]  # MIIP SS
rj_r_s = 6 + 18 * N_S + 1   # linha RJ inicio (1-indexed)
rj_r_e = 6 + 19 * N_S        # linha RJ fim
rj_c_s = 4 + 18 * N_S + 1   # coluna RJ inicio
rj_c_e = 4 + 19 * N_S        # coluna RJ fim
print(f'\n=== MIIP SS: Bloco RJ ===')
print(f'Linhas: {rj_r_s} - {rj_r_e} | Colunas: {rj_c_s} - {rj_c_e}')
print(f'Dim: {rj_r_e - rj_r_s + 1} x {rj_c_e - rj_c_s + 1}')

# Ler 5 primeiras linhas, com identificadores
print('Primeiras 5 linhas (colunas 1-8):')
for i, row in enumerate(ws.iter_rows(min_row=rj_r_s, max_row=rj_r_s+4,
                                      min_col=1, max_col=8, values_only=True)):
    vals = [str(v)[:22] if v is not None else '' for v in row]
    print(f'  R{rj_r_s+i}: {vals}')

# Verificar soma de uma linha inteira do bloco (todos os destinos)
print('\nSoma dos fluxos totais de S01 do RJ para todos os destinos:')
TOTAL_COLS = ws.max_column
print(f'  Total de colunas existentes: {TOTAL_COLS}')

# Ler linha RJ S01 inteira (para ver total de saídas)
row_data = []
for row in ws.iter_rows(min_row=rj_r_s, max_row=rj_r_s,
                         min_col=1, max_col=min(2000, TOTAL_COLS), values_only=True):
    for v in row:
        try:
            row_data.append(float(v))
        except:
            row_data.append(0.0)

# Valores de dados (a partir da coluna 5)
data_vals = np.array(row_data[4:])
print(f'  Soma total de saidas S01 RJ: {data_vals.sum():.2f}')
print(f'  Saidas para fora do bloco RJ: {data_vals[:18*N_S].sum() + data_vals[19*N_S:].sum():.2f}')
print(f'  Saidas dentro do bloco RJ (diagonal): {data_vals[18*N_S:19*N_S].sum():.2f}')

# ── Verificar aba Produção para VBP ─────────────────────────────────────────
ws2 = wb[names[4]]  # Produção
print(f'\n=== ABA PRODUCAO ===')
rj_row_s2 = 6 + 18 * N_S + 1
rj_row_e2 = 6 + 19 * N_S
print(f'Bloco RJ: linhas {rj_row_s2} - {rj_row_e2}')
print(f'Colunas totais: {ws2.max_column}')

# Verificar primeiras 3 linhas RJ
print('Primeiras 3 linhas (colunas 1-8):')
for i, row in enumerate(ws2.iter_rows(min_row=rj_row_s2, max_row=rj_row_s2+2,
                                       min_col=1, max_col=8, values_only=True)):
    vals = [str(v)[:25] if v is not None else '' for v in row]
    print(f'  R{rj_row_s2+i}: {vals}')

# Calcular VBP de cada setor RJ = soma de toda a linha de producao
print('\nCalculando VBP por setor RJ (soma de toda a linha de producao)...')
vbp_rj = np.zeros(N_S)
for i, row in enumerate(ws2.iter_rows(min_row=rj_row_s2, max_row=rj_row_e2,
                                       min_col=5, values_only=True)):
    total = 0.0
    for v in row:
        try:
            total += float(v)
        except:
            pass
    vbp_rj[i] = total

print(f'VBP total RJ: R$ {vbp_rj.sum():,.0f} Mi')
print(f'Setores com VBP > 0: {(vbp_rj > 0).sum()} de {N_S}')
print(f'Setores com VBP zero: {(vbp_rj == 0).sum()}')
print('Top 10 setores por VBP:')
idx_sorted = np.argsort(vbp_rj)[::-1]
for k in range(10):
    i = idx_sorted[k]
    print(f'  S{i+1:02d}: R$ {vbp_rj[i]:,.1f} Mi')

wb.close()

# Salvar VBP para uso posterior
np.save(r'C:\Users\jonat\Documents\MIP e CGE\output\crosswalk\vbp_iioas_rj.npy', vbp_rj)
print(f'\nVBP salvo: output/crosswalk/vbp_iioas_rj.npy')
