from openpyxl import load_workbook

FILE_PATH = r"C:\Users\jonat\Documents\MIP e CGE\data\2026.01.08.xls"

def inspect_confaz_xlsx():
    """Memory-efficient XLSX inspection using read_only mode"""
    
    print("="*70)
    print("CONFAZ ICMS FILE STRUCTURE")
    print("="*70)
    
    try:
        # Load in read_only mode (memory efficient)
        wb = load_workbook(FILE_PATH, read_only=True, data_only=True)
        
        print(f"\nSheets: {wb.sheetnames}")
        
        # Get first sheet
        sheet = wb[wb.sheetnames[0]]
        
        print(f"\n[First 25 rows - showing all non-empty columns]:")
        print("-"*70)
        
        row_count = 0
        for row in sheet.iter_rows(min_row=1, max_row=25, values_only=True):
            # Show only non-empty cells
            non_empty = [(i, v) for i, v in enumerate(row) if v is not None and str(v).strip() != '']
            if non_empty:
                row_str = ' | '.join([f"Col{chr(65+i)}={str(v)[:40]}" for i, v in non_empty[:8]])
                print(f"Row {row_count+1}: {row_str}")
            row_count += 1
        
        # Focus on columns D and E (user mentioned these are Year and Month)
        print(f"\n[Columns D and E specifically - First 25 rows]:")
        print("-"*70)
        row_count = 0
        for row in sheet.iter_rows(min_row=1, max_row=25, min_col=4, max_col=5, values_only=True):
            col_d, col_e = row[0] if len(row) > 0 else None, row[1] if len(row) > 1 else None
            if col_d or col_e:
                print(f"Row {row_count+1}: D(Year)={col_d} | E(Month)={col_e}")
            row_count += 1
        
        wb.close()
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    inspect_confaz_xlsx()
