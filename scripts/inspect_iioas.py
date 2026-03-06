"""
Script detalhado para extrair toda a estrutura do IIOAS_BRUF_2019
"""
import openpyxl, sys

FILE = r'C:\Users\jonat\Documents\MIP e CGE\data\IIOAS_BRUF_2019.xlsx'
# Force UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

print("Carregando arquivo...")
wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
names = wb.sheetnames
print("Abas:", names)

# ---- ABA CREDITOS ----
ws = wb[names[0]]  # Créditos
print("\n=== ABA: CREDITOS ===")
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i >= 30: break
    vals = [str(v)[:80] if v else '' for v in row[:5]]
    if any(vals):
        print(f"  {vals}")

# ---- ABA REGIOES ----
ws = wb[names[1]]  # Regiões
print("\n=== ABA: REGIOES ===")
print(f"  Dimensoes: {ws.max_row} x {ws.max_column}")
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i >= 2 and (row[0] is None and row[1] is None): continue
    vals = [str(v) if v else '' for v in row[:3]]
    if any(vals):
        print(f"  L{i+1:02d}: {vals}")

# ---- ABA SETORES ----
ws = wb[names[2]]  # Setores
print("\n=== ABA: SETORES ===")
print(f"  Dimensoes: {ws.max_row} x {ws.max_column}")
all_rows = list(ws.iter_rows(values_only=True))
print("  TODOS OS SETORES:")
for i, row in enumerate(all_rows):
    if i < 3: continue
    if row[0] is None: continue
    print(f"    [{row[0]}] {row[1]}")

# ---- ABA PRODUTOS ----
ws = wb[names[3]]  # Produtos
print("\n=== ABA: PRODUTOS ===")
print(f"  Dimensoes: {ws.max_row} x {ws.max_column}")
all_rows = list(ws.iter_rows(values_only=True))
print("  TODOS OS PRODUTOS:")
for i, row in enumerate(all_rows):
    if i < 3: continue
    if row[0] is None: continue
    print(f"    [{row[0]}] {row[1]}")

# ---- ABA PRODUCAO ----
ws = wb[names[4]]  # Produção
print("\n=== ABA: PRODUCAO ===")
print(f"  Dimensoes declaradas: {ws.max_row} x {ws.max_column}")
print("  Titulo (L1):")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True)):
    vals = [str(v)[:50] if v else '' for v in list(row)[:10]]
    print(f"    L{i+1}: {vals}")

# ---- ABA MIIP PS ----
ws = wb[names[5]]  # MIIP PS
print("\n=== ABA: MIIP PS ===")
print(f"  Dimensoes declaradas: {ws.max_row} x {ws.max_column}")
print("  Cabecalho (primeiras 6 linhas, primeiras 8 colunas):")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True)):
    vals = [str(v)[:50] if v else '' for v in list(row)[:8]]
    print(f"    L{i+1}: {vals}")

# ---- ABA MIIP SS ----
ws = wb[names[6]]  # MIIP SS
print("\n=== ABA: MIIP SS ===")
print(f"  Dimensoes declaradas: {ws.max_row} x {ws.max_column}")
print("  Cabecalho (primeiras 6 linhas, primeiras 8 colunas):")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True)):
    vals = [str(v)[:50] if v else '' for v in list(row)[:8]]
    print(f"    L{i+1}: {vals}")

wb.close()
print("\n\nFinalizado.")
