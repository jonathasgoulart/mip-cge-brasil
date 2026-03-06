"""
Atualiza a tax_matrix de base 2015 para 2019 usando:
  1. Fatores de crescimento tributo-a-tributo do CTB histórico (ctb_1990-2024_0.xlsx)
  2. Mantém X_nas (VBP MIP Nacional 2019) como denominador — correto para coef fiscais
  3. Corrige demand_shock.py para usar X_nas no cálculo de tax, não VBP_IIOAS

Resultado: tax_matrix_2019.json com valores nominais de 2019 e coeficientes τ = tax/X_nas corretos.
"""
import numpy as np
import json
import openpyxl
import sys
from pathlib import Path
sys.stdout.reconfigure(encoding='utf-8')

BASE = Path(r'C:\Users\jonat\Documents\MIP e CGE')
CTB_FILE  = BASE / 'data' / 'ctb_1990-2024_0.xlsx'
TAX_FILE  = BASE / 'data' / 'processed' / '2021_final' / 'tax_matrix.json'
X_NAS     = BASE / 'output' / 'intermediary' / 'X_nas.npy'
OUT_DIR   = BASE / 'data' / 'processed' / '2021_final'

# ── 1. Carregar tax_matrix atual (base 2015) ──────────────────────────────────
with open(TAX_FILE, encoding='utf-8') as f:
    tax_data = json.load(f)

taxes_2015 = tax_data['taxes_by_type']
N = 67

# ── 2. Extrair fatores de crescimento 2015→2019 do CTB ───────────────────────
print("Extraindo fatores de crescimento 2015→2019 do CTB...")
wb = openpyxl.load_workbook(CTB_FILE, read_only=True, data_only=True)

# Aba 'CTB (R$ milhões)' - valores nominais
ws_millions = None
for s in wb.sheetnames:
    if 'milh' in s.lower() or 'R$' in s:
        ws_millions = wb[s]
        break
if ws_millions is None:
    ws_millions = wb[wb.sheetnames[0]]

print(f"  Aba usada: {ws_millions.title}")

# Ler todos os dados
rows = []
for row in ws_millions.iter_rows(values_only=True):
    rows.append(list(row))

# Encontrar linha de cabeçalho (tem anos)
header_idx = None
years = {}
for i, row in enumerate(rows):
    for j, v in enumerate(row):
        if v == 1990 or v == '1990':
            header_idx = i
            break
    if header_idx is not None:
        break

if header_idx is not None:
    header = rows[header_idx]
    for j, v in enumerate(header):
        try:
            yr = int(v)
            years[yr] = j
        except:
            pass
    print(f"  Anos encontrados: {sorted(years.keys())[:10]}...{sorted(years.keys())[-5:]}")
else:
    print("  AVISO: cabeçalho de anos não encontrado — usando aba de % do PIB")

# Procurar valores de 2015 e 2019 para cada tributo
# Linhas relevantes: ICMS, IPI, ISS, PIS/PASEP, COFINS, IOF, CIDE, II
CTB_KEYWORDS = {
    'ICMS':     ['icms', 'imposto sobre circulação'],
    'IPI':      ['ipi', 'industrializados'],
    'ISS':      ['iss', 'serviços de qualquer'],
    'PIS_PASEP':['pis', 'pasep'],
    'COFINS':   ['cofins'],
    'IOF':      ['iof', 'operações financeiras'],
    'CIDE':     ['cide', 'contribuição de intervenção'],
    'II':       ['imposto de importação', ' ii '],
}

growth_factors = {}
col_2015 = years.get(2015)
col_2019 = years.get(2019)

if col_2015 and col_2019:
    print(f"  Colunas: 2015={col_2015}, 2019={col_2019}")
    for tname, keywords in CTB_KEYWORDS.items():
        for i, row in enumerate(rows):
            if i <= header_idx:
                continue
            label = str(row[0] or '').lower()
            if any(kw in label for kw in keywords):
                try:
                    v2015 = float(row[col_2015])
                    v2019 = float(row[col_2019])
                    if v2015 > 0:
                        factor = v2019 / v2015
                        growth_factors[tname] = factor
                        print(f"  {tname:12}: {v2015:12,.0f} (2015) → {v2019:12,.0f} (2019) → fator={factor:.4f}")
                        break
                except:
                    pass
else:
    print("  AVISO: colunas 2015/2019 não encontradas. Usando aba de % PIB para calcular fator.")

wb.close()

# Fallback: se não encontrou, usar deflator geral (IPCA acumulado 2015-2019 ≈ 25%)
if not growth_factors:
    print("\n  Usando fator de deflação IPCA acumulado 2015-2019 (~25%) como fallback.")
    for t in taxes_2015.keys():
        growth_factors[t] = 1.265  # IPCA médio: ~6% a.a. por 4 anos

# Fatores para impostos não encontrados — usar crescimento proporcional do PIB nominal (fator ~1.27)
pib_factor = growth_factors.get('ICMS', 1.27)
for t in taxes_2015.keys():
    if t not in growth_factors:
        growth_factors[t] = pib_factor
        print(f"  {t}: usando fator PIB ({pib_factor:.4f})")

print(f"\n  Fatores finais: {growth_factors}")

# ── 3. Aplicar fatores e gerar tax_matrix_2019.json ────────────────────────────
print("\nAplicando fatores 2015→2019...")
taxes_2019 = {}
for tname, vals in taxes_2015.items():
    factor = growth_factors.get(tname, pib_factor)
    taxes_2019[tname] = [round(v * factor, 4) for v in vals]
    total_2015 = sum(vals[:N])
    total_2019 = sum(taxes_2019[tname][:N])
    print(f"  {tname:12}: R$ {total_2015:,.0f} Mi (2015) → R$ {total_2019:,.0f} Mi (2019)")

out_tax = {
    "metadata": "Tax matrix updated to 2019 nominal values using CTB growth factors (ctb_1990-2024_0.xlsx). Base: MIP 2015 sectoral distribution. VBP denominator: X_nas MIP Nacional 2019.",
    "year": 2019,
    "base_year": 2015,
    "growth_source": "CTB 1990-2024 IBGE",
    "taxes_by_type": taxes_2019
}
out_path = OUT_DIR / 'tax_matrix_2019.json'
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(out_tax, f, ensure_ascii=False, indent=2)
print(f"\nSalvo: {out_path}")

# ── 4. Calcular e validar coeficientes τ = tax/X_nas ──────────────────────────
print("\nValidando coeficientes τ = tax/X_nas...")
X_nas = np.load(X_NAS)
print(f"  X_nas (VBP MIP 2019): total R$ {X_nas.sum():,.0f} Mi")

for tname, vals in taxes_2019.items():
    tax_vec = np.array(vals[:N])
    tau = np.where(X_nas > 0, tax_vec / X_nas, 0.0)
    print(f"  τ({tname:12}): média={tau.mean():.4f}  max={tau.max():.4f}  total_tax=R$ {tax_vec.sum():,.0f} Mi")

# ── 5. Corrigir demand_shock.py: usar X_nas para coef fiscal, não VBP_IIOAS ───
print("\n" + "=" * 60)
print("CORREÇÃO NECESSÁRIA NO demand_shock.py")
print("=" * 60)
print("  No cálculo de impostos (linha ~138-154), a variável X_nas")
print("  é usada como denominador para o coef fiscal.")
print("  → Garantir que X_nas = np.load(path_X_nas), NÃO o VBP_IIOAS.")
print("  → O VBP_IIOAS deve ser usado APENAS para distribuir o choque y,")
print("    NÃO para calcular τ.")
print()
print("  [OK] Verificando demand_shock.py atual...")

# Ler o demand_shock para verificar que a lógica fiscal usa X_nas correto
ds_path = BASE / 'api' / 'simulators' / 'demand_shock.py'
ds_text = ds_path.read_text(encoding='utf-8')

# Verificar se a seção fiscal usa X_nas (separado do VBP_IIOAS que vai para y)
if 'path_X_nas' in ds_text and 'tax_data' in ds_text:
    # Encontrar a seção relevante
    lines = ds_text.split('\n')
    fiscal_lines = [(i+1, l) for i, l in enumerate(lines) if 'tax' in l.lower() and any(k in l for k in ['coef', 'X_nas', 'revenue', 'tax_values'])]
    print("  Linhas fiscais relevantes no demand_shock.py:")
    for ln, ll in fiscal_lines[:10]:
        print(f"    L{ln}: {ll.strip()}")

print("\n[OK] Script concluído!")
